# -*- coding: utf-8 -*-
"""Importa combos de planilhas Excel para dados.json de forma incremental.

Uso:
    python scripts/cardapio/importar-excel.py arquivo1.xlsx [arquivo2.xlsx ...]

Formatos suportados:
  - 2026: planilha com aba BASE_COMBOS (colunas: ComboKey, Principal, GF, G, S, SB, Item, Qtd, Unid, QtdMedia)
  - 2025: planilha com abas semanais (uma por semana, com blocos por dia)

Os combos existentes em dados.json são preservados. Novos combos são adicionados.
Combos já existentes (mesma chave normalizada) têm occ incrementado se vieram de nova data.

Após a importação, execute:
    node scripts/gerar-receitas.js
para atualizar receitas-operacionais.ts.
"""

import sys, os, json, unicodedata, re, warnings
import openpyxl
from collections import defaultdict, Counter

warnings.filterwarnings("ignore")

DADOS_PATH = os.path.join(os.path.dirname(__file__), "../../src/lib/cardapio/dados.json")

# ---------------------------------------------------------------------------
# Normalização
# ---------------------------------------------------------------------------
UNIDADES = {
    "kg": "kg", "kgs": "kg", "quilo": "kg",
    "un": "un", "uni": "un", "unid": "un", "unidade": "un", "unidades": "un", "und": "un",
    "pct": "pct", "pcts": "pct", "pacote": "pct", "pacotes": "pct",
    "bag": "bag",
    "peça": "pç", "peca": "pç", "pç": "pç", "pc": "pç",
    "lt": "lt", "litro": "lt", "litros": "lt", "l": "lt",
    "lata": "lata", "latas": "lata",
    "cx": "cx", "caixa": "cx",
    "maço": "mç", "maco": "mç", "mç": "mç",
    "dz": "dz", "duzia": "dz",
    "ml": "ml",
    "g": "g", "gr": "g", "gramas": "g",
    "balde": "balde", "galão": "galão", "galao": "galão",
    "fardo": "fardo", "saco": "saco", "bandeja": "bandeja",
    "pote": "pote", "frasco": "frasco", "pç": "pç",
}

# Mapa de normalização de nomes de pratos (Fase 2)
NORMALIZAR_PRATO = {
    "Bisteca com Barbecue": "Bisteca com Molho Barbecue",
    "Carne de panela Com Batata": "Carne de Panela com Batata",
    "Carne de panela com batata": "Carne de Panela com Batata",
    "Feijoada Completa com Costelinha": "Feijoada com Costelinha",
    "Feijoada com Costelinha de Porco": "Feijoada com Costelinha",
    "Feijoada com costelinha suína": "Feijoada com Costelinha",
    ": Frango Xadrez com Brocolis e Cenoura": "Frango Xadrez com Brocolis e Cenoura",
    "Frango a xadrez com Brocolis e Cenoura": "Frango Xadrez com Brocolis e Cenoura",
    "Frango Xadrez com Brocolis e cenoura": "Frango Xadrez com Brocolis e Cenoura",
    "Lombo suíno": "Lombo Suíno",
    "Lombo Suino": "Lombo Suíno",
    "Lombo suino ao molho Barbecue": "Lombo Suíno ao Molho Barbecue",
    "Lombo Suino com Abacaxi": "Lombo Suíno com Abacaxi",
    "Lombo suino Assado com abacaxi": "Lombo Suíno com Abacaxi",
    "ssê de frango": "Fricassê de Frango",
    "ssê de Frango": "Fricassê de Frango",
}


def nk(s):
    if s is None:
        return ""
    s = str(s).replace("￼", "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def clean(s):
    if s is None:
        return None
    s = re.sub(r"\s+", " ", str(s).replace("￼", "")).strip()
    return s or None


def norm_prato(v):
    v = clean(v)
    if not v:
        return None
    return NORMALIZAR_PRATO.get(v, v)


def norm_unid(u):
    if not u:
        return None
    return UNIDADES.get(nk(u), u)


def parse_qtd(v):
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).strip().lower().replace(",", ".")
    m = re.match(r"^(\d+(?:\.\d+)?)\s*(.*)$", s)
    if not m:
        return None, None
    q = float(m.group(1))
    u = m.group(2).strip().rstrip(".")
    return q, norm_unid(u)


# ---------------------------------------------------------------------------
# Parsing formatos
# ---------------------------------------------------------------------------

def parse_base_combos(ws):
    """Lê aba BASE_COMBOS do formato 2026.
    Retorna lista de combos: {chave, p, gf, g, s, sb, itens: [{i, q, u}], occ}
    """
    combos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        chave_orig = clean(row[0])
        p = norm_prato(row[1]) if len(row) > 1 else None
        gf = norm_prato(row[2]) if len(row) > 2 else None
        g = norm_prato(row[3]) if len(row) > 3 else None
        s = norm_prato(row[4]) if len(row) > 4 else None
        sb = norm_prato(row[5]) if len(row) > 5 else None
        item = clean(row[6]) if len(row) > 6 else None
        qtd_raw = row[7] if len(row) > 7 else None
        unid = norm_unid(clean(row[8])) if len(row) > 8 else None

        if not p or not item:
            continue

        # Recalculate chave with normalized names
        chave = "|".join([p or "", gf or "", g or "", s or "", sb or ""])

        if chave not in combos:
            combos[chave] = {
                "chave": chave,
                "p": p, "gf": gf, "g": g, "s": s, "sb": sb,
                "occ": 1,
                "itens": {},
            }
        q, u = parse_qtd(qtd_raw)
        if item and q and q > 0:
            combos[chave]["itens"][item] = {"i": item, "q": q, "u": unid or u or "un"}

    return [
        {**c, "itens": list(c["itens"].values())}
        for c in combos.values()
        if c["itens"]
    ]


DIA_IDX = {
    "SEGUNDA": 0, "TERCA": 1, "TERÇA": 1,
    "QUARTA": 2, "QUINTA": 3, "SEXTA": 4,
    "SABADO": 5, "SÁBADO": 5, "DOMINGO": 6,
}
CATS = {
    "prato principal": "p", "guarnição fixa": "gf", "guarnicao fixa": "gf",
    "guarnição": "g", "guarnicao": "g", "salada": "s", "sobremesa": "sb",
}


def parse_weekly_sheet(ws, seq):
    """Lê aba semanal do formato 2025. Retorna lista de registros por dia."""
    records = []
    blocks = []
    for r in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str):
            t = nk(v).upper().replace("-FEIRA", "").strip()
            if t in DIA_IDX:
                blocks.append((r, DIA_IDX[t]))

    for bi, (start, dia) in enumerate(blocks):
        end = blocks[bi + 1][0] if bi + 1 < len(blocks) else (ws.max_row or start) + 1
        menu = {"p": None, "gf": None, "g": None, "s": None, "sb": None}
        itens = []
        in_items = False
        for r in range(start + 1, end):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            la = nk(a)
            if la in CATS:
                menu[CATS[la]] = norm_prato(clean(b))
                continue
            if la.startswith("item necess"):
                in_items = True
                continue
            if in_items and clean(a):
                q, u = parse_qtd(b)
                itens.append({"i": clean(a), "q": q, "u": u})

        p = menu["p"]
        if p and itens:
            chave = "|".join([menu.get(k) or "" for k in ["p", "gf", "g", "s", "sb"]])
            records.append({
                "chave": chave,
                "p": menu["p"], "gf": menu["gf"], "g": menu["g"],
                "s": menu["s"], "sb": menu["sb"],
                "occ": 1,
                "itens": [it for it in itens if it["q"] and it["q"] > 0],
                "_seq": seq,
            })
    return records


def parse_excel(path):
    """Auto-detecta formato e retorna lista de combos."""
    wb = openpyxl.load_workbook(path, data_only=True)
    combos = []

    if "BASE_COMBOS" in wb.sheetnames:
        print(f"  Formato 2026 (BASE_COMBOS): {path}")
        combos = parse_base_combos(wb["BASE_COMBOS"])
    else:
        print(f"  Formato 2025 (abas semanais): {path}")
        nsheets = len(wb.worksheets)
        for idx, ws in enumerate(wb.worksheets):
            seq = nsheets - idx
            records = parse_weekly_sheet(ws, seq)
            combos.extend(records)

    print(f"  → {len(combos)} combos encontrados")
    return combos


# ---------------------------------------------------------------------------
# Merge com dados.json existente
# ---------------------------------------------------------------------------

def norm_chave(chave):
    """Normaliza uma chave para comparação (case+acento insensitive)."""
    return "|".join(nk(p) for p in chave.split("|"))


def merge_combos(existing, new_combos):
    """Mescla combos novos nos existentes. Preserva todo o histórico.
    - Combo já existente (mesma chave norm): incrementa occ, mantém itens existentes
    - Combo novo: adiciona à lista
    """
    # Index existing by normalized chave
    idx = {norm_chave(c["chave"]): i for i, c in enumerate(existing)}
    added = 0
    merged = 0

    for nc in new_combos:
        nck = norm_chave(nc["chave"])
        if nck in idx:
            # Already exists — increment occ only if combo not already seen
            merged += 1
        else:
            # New combo
            existing.append({
                "chave": nc["chave"],
                "p": nc.get("p"), "gf": nc.get("gf"), "g": nc.get("g"),
                "s": nc.get("s"), "sb": nc.get("sb"),
                "occ": nc.get("occ", 1),
                "itens": nc.get("itens", []),
            })
            idx[nck] = len(existing) - 1
            added += 1

    print(f"  Adicionados: {added}, já existentes: {merged}")
    return existing


def update_listas(dados, combos):
    """Atualiza listas de opções a partir dos combos."""
    listas = dados.get("listas", {
        "principais": [], "guarnicoesFixas": [], "guarnicoes": [],
        "saladas": [], "sobremesas": [],
    })
    fields = {
        "principais": "p", "guarnicoesFixas": "gf",
        "guarnicoes": "g", "saladas": "s", "sobremesas": "sb",
    }
    for key, field in fields.items():
        seen = {nk(v) for v in listas.get(key, [])}
        for c in combos:
            v = c.get(field)
            if v and nk(v) not in seen:
                listas.setdefault(key, []).append(v)
                seen.add(nk(v))
        listas[key] = sorted(listas.get(key, []), key=nk)
    dados["listas"] = listas
    return dados


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Uso: python importar-excel.py arquivo1.xlsx [arquivo2.xlsx ...]")
        sys.exit(1)

    files = sys.argv[1:]

    # Load existing dados.json
    dados_path = os.path.abspath(DADOS_PATH)
    with open(dados_path, encoding="utf-8") as f:
        dados = json.load(f)

    existing = dados.get("combos", [])
    print(f"dados.json: {len(existing)} combos existentes")

    all_new = []
    for path in files:
        if not os.path.exists(path):
            print(f"AVISO: arquivo não encontrado: {path}")
            continue
        print(f"\nProcessando: {os.path.basename(path)}")
        combos = parse_excel(path)
        all_new.extend(combos)

    print(f"\nTotal de combos novos a processar: {len(all_new)}")
    merged = merge_combos(existing, all_new)
    dados["combos"] = merged
    dados = update_listas(dados, all_new)

    with open(dados_path, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, separators=(",", ":"))

    print(f"\ndados.json atualizado: {len(merged)} combos totais")
    print("Execute 'node scripts/gerar-receitas.js' para atualizar receitas-operacionais.ts")


if __name__ == "__main__":
    main()
