# -*- coding: utf-8 -*-
"""Consolida o banco de dados do cardápio a partir de:
   - Pedido 2025 (47 semanas brutas: cardápio + itens + qtd)
   - Abril/Maio/Junho 2026 (abas Semana com valores calculados + BASE_MAPAS)
   Gera src/lib/cardapio/dados.json
"""
import openpyxl, unicodedata, re, json, warnings
from collections import Counter, defaultdict

warnings.filterwarnings("ignore")
U = "/root/.claude/uploads/28a7abfe-f16f-54cd-9c25-3a5680d3125b/"
F2025 = U + "76d86e48-Planilha_Tata_House_Pedido_2025.xlsx"
F2026 = [U + "f1a19944-TATA__House_Abril_2026.xlsx",
         U + "efc176e0-TAT__House_Maio_.xlsx",
         U + "01843faf-TAT__House_Junho.xlsx"]

def nk(s):
    if s is None: return ""
    s = str(s).replace("￼","")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+"," ",s).strip().lower()

def clean(s):
    if s is None: return None
    s = re.sub(r"\s+"," ",str(s).replace("￼","")).strip()
    return s or None

UNIDADES = {"kg":"kg","kgs":"kg","quilo":"kg","un":"un","uni":"un","unid":"un","unidade":"un",
            "unidades":"un","und":"un","pct":"pct","pcts":"pct","pacote":"pct","pacotes":"pct",
            "bag":"bag","peça":"pç","peca":"pç","pç":"pç","pc":"pç","lt":"lt","litro":"lt","litros":"lt",
            "l":"lt","lata":"lata","latas":"lata","cx":"cx","caixa":"cx","maço":"mç","maco":"mç","mç":"mç",
            "dz":"dz","duzia":"dz","ml":"ml","g":"g","gr":"g","gramas":"g","balde":"balde","galão":"galão",
            "galao":"galão","fardo":"fardo","saco":"saco","bandeja":"bandeja","pote":"pote","frasco":"frasco"}

def parse_qtd(v):
    """'20 kg' -> (20.0,'kg'); 7 -> (7.0,None)"""
    if v is None: return None
    if isinstance(v,(int,float)): return (float(v), None)
    s = str(v).strip().lower().replace(",",".")
    m = re.match(r"^(\d+(?:\.\d+)?)\s*(.*)$", s)
    if not m: return None
    q = float(m.group(1))
    u = m.group(2).strip().rstrip(".")
    u = UNIDADES.get(nk(u), u if u else None)
    return (q, u or None)

DIA_IDX = {"SEGUNDA":0,"TERCA":1,"TERÇA":1,"QUARTA":2,"QUINTA":3,"SEXTA":4,"SABADO":5,"SÁBADO":5,"DOMINGO":6}
CATS = {"prato principal":"p","guarnição fixa":"gf","guarnicao fixa":"gf","guarnição":"g",
        "guarnicao":"g","salada":"s","sobremesa":"sb"}

records = []   # {seq, dia, menu:{p,gf,g,s,sb}, itens:[(nome,q,u)]}

# ---------- 1. Pedido 2025 (abas em ordem: mais recente primeiro) ----------
wb = openpyxl.load_workbook(F2025, data_only=True)
nsheets = len(wb.worksheets)
for idx, ws in enumerate(wb.worksheets):
    seq = nsheets - idx          # maior = mais recente
    # localizar blocos de dia
    blocks = []
    for r in range(1, (ws.max_row or 0)+1):
        v = ws.cell(r,1).value
        if isinstance(v,str):
            t = nk(v).upper().replace("-FEIRA","").strip()
            if t in DIA_IDX: blocks.append((r, DIA_IDX[t]))
    for bi,(start, dia) in enumerate(blocks):
        end = blocks[bi+1][0] if bi+1 < len(blocks) else (ws.max_row or start)+1
        menu = {"p":None,"gf":None,"g":None,"s":None,"sb":None}
        itens = []
        in_items = False
        for r in range(start+1, end):
            a = ws.cell(r,1).value; b = ws.cell(r,2).value
            la = nk(a)
            if la in CATS:
                menu[CATS[la]] = clean(b); continue
            if la.startswith("item necess"):
                in_items = True; continue
            if in_items and clean(a):
                pq = parse_qtd(b)
                itens.append((clean(a), pq[0] if pq else None, pq[1] if pq else None))
        if menu["p"] and itens:
            records.append({"seq":seq,"dia":dia,"menu":menu,"itens":itens})
print(f"Pedido 2025: {len(records)} dias com cardápio+itens")

# ---------- 2. Meses 2026: abas Semana (valores em cache) ----------
mapas_rows = []   # (tipo, opcao, item, unid, qtd, occ)
listas_extra = defaultdict(list)
STARTS = [18,42,66,90,114,138,162]
base_seq = nsheets + 1
for mi, f in enumerate(F2026):
    wb = openpyxl.load_workbook(f, data_only=True)
    # BASE_MAPAS
    if "BASE_MAPAS" in wb.sheetnames:
        bm = wb["BASE_MAPAS"]
        for row in bm.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1] and row[2]:
                mapas_rows.append((str(row[0]).strip().lower(), clean(row[1]), clean(row[2]),
                                   row[3], float(row[4] or 0), float(row[6] or 1)))
    # LISTAS
    if "LISTAS" in wb.sheetnames:
        ls = wb["LISTAS"]
        for ci,key in enumerate(["principais","guarnicoesFixas","guarnicoes","saladas","sobremesas"],start=1):
            for r in range(2, (ls.max_row or 1)+1):
                v = clean(ls.cell(r,ci).value)
                if v: listas_extra[key].append(v)
    # Semanas
    wnames = [n for n in wb.sheetnames if n.strip().lower().startswith("semana")]
    for wi, wn in enumerate(sorted(wnames)):
        ws = wb[wn]
        seq = base_seq + mi*10 + wi
        for d in range(7):
            menu = {"p":clean(ws.cell(6+d,3).value),"gf":clean(ws.cell(6+d,4).value),
                    "g":clean(ws.cell(6+d,5).value),"s":clean(ws.cell(6+d,6).value),
                    "sb":clean(ws.cell(6+d,7).value)}
            s = STARTS[d]
            itens = []
            for r in range(s, s+21):
                nome = clean(ws.cell(r,1).value)
                if not nome or nome.startswith("="): continue
                unid = clean(ws.cell(r,2).value)
                qtd  = ws.cell(r,7).value   # Qtd final
                if not isinstance(qtd,(int,float)): qtd = ws.cell(r,5).value
                if not isinstance(qtd,(int,float)) or qtd <= 0: continue
                itens.append((nome, float(qtd), UNIDADES.get(nk(unid), unid) if unid else None))
            if menu["p"] and itens:
                records.append({"seq":seq,"dia":d,"menu":menu,"itens":itens})
print(f"Total com 2026: {len(records)} dias")

# ---------- 3. canonicalização global ----------
vc = Counter()
for rec in records:
    for v in rec["menu"].values():
        if v: vc[v] += 1
    for nome,_,_ in rec["itens"]: vc[nome] += 1
for _,op,it,_,_,_ in mapas_rows:
    vc[op] += 1; vc[it] += 1
for key, vals in listas_extra.items():
    for v in vals: vc[v] += 1

groups = defaultdict(list)
for v,c in vc.items(): groups[nk(v)].append((c,v))
canon = {}
for k,lst in groups.items():
    lst.sort(key=lambda t:(-t[0],t[1]))
    canon[k] = lst[0][1]
def C(v):
    v = clean(v)
    return canon.get(nk(v), v) if v else None

# ---------- 4. itens: unidade + frequência ----------
item_units = defaultdict(Counter); item_freq = Counter()
for rec in records:
    for nome,q,u in rec["itens"]:
        cn = C(nome); item_freq[cn] += 1
        if u: item_units[cn][u] += 1
for _,_,it,u,_,occ in mapas_rows:
    cn = C(it)
    if u: item_units[cn][nk(u) and UNIDADES.get(nk(u),u)] += 1
itens_out = []
for nome,f in item_freq.most_common():
    u = item_units[nome].most_common(1)
    itens_out.append({"n":nome,"u":(u[0][0] if u else "un") or "un","f":f})
# itens que só existem nos mapas
for cn in {C(m[2]) for m in mapas_rows}:
    if cn not in item_freq:
        u = item_units[cn].most_common(1)
        itens_out.append({"n":cn,"u":(u[0][0] if u else "un") or "un","f":1})

# ---------- 5. combos: tupla canônica -> itens (registro mais recente) ----------
combo_best = {}   # key -> (seq, rec)
combo_occ = Counter()
for rec in records:
    m = {k:(C(v) or "") for k,v in rec["menu"].items()}
    key = "|".join([m["p"],m["gf"],m["g"],m["s"],m["sb"]])
    combo_occ[key] += 1
    if key not in combo_best or rec["seq"] >= combo_best[key][0]:
        combo_best[key] = (rec["seq"], rec, m)
combos_out = []
for key,(seq,rec,m) in combo_best.items():
    agg = {}
    for nome,q,u in rec["itens"]:
        cn = C(nome)
        if q is None: continue
        unit = u or (item_units[cn].most_common(1)[0][0] if item_units[cn] else "un")
        if cn in agg: agg[cn] = (max(agg[cn][0],q), agg[cn][1])
        else: agg[cn] = (q, unit)
    if not agg: continue
    combos_out.append({"chave":key,"p":m["p"] or None,"gf":m["gf"] or None,"g":m["g"] or None,
                       "s":m["s"] or None,"sb":m["sb"] or None,"occ":combo_occ[key],
                       "itens":[{"i":k,"q":v[0],"u":v[1]} for k,v in agg.items()]})
print(f"Combos únicos: {len(combos_out)}")

# ---------- 6. mapas agregados ----------
mapas_g = {}
for tipo,op,it,u,q,occ in mapas_rows:
    gk = (tipo, nk(op), nk(it))
    g = mapas_g.get(gk)
    if g and occ <= g["occ"]:
        g["occ"] += occ
    else:
        prev = g["occ"] if g else 0
        mapas_g[gk] = {"tipo":tipo,"op":C(op),"it":C(it),
                       "u":UNIDADES.get(nk(u),u) if u else None,"q":q,"occ":occ+prev}
mapas_out = defaultdict(list)
for g in mapas_g.values():
    mapas_out[(g["tipo"],g["op"])].append({"i":g["it"],"q":g["q"],"u":g["u"]})
mapas_final = [{"tipo":t,"op":o,"itens":v} for (t,o),v in mapas_out.items()]
print(f"Mapas (tipo+opção): {len(mapas_final)}")

# ---------- 7. listas de opções ----------
listas = {k: [] for k in ["principais","guarnicoesFixas","guarnicoes","saladas","sobremesas"]}
fontes = {"principais":("p",), "guarnicoesFixas":("gf",), "guarnicoes":("g",),
          "saladas":("s",), "sobremesas":("sb",)}
for key,(mk,) in fontes.items():
    seen = set(); out = []
    vals = [C(v) for v in listas_extra[key]]
    vals += [C(rec["menu"][mk]) for rec in records]
    for v in vals:
        if v and nk(v) not in seen:
            seen.add(nk(v)); out.append(v)
    listas[key] = sorted(out, key=nk)
print("Listas:", {k:len(v) for k,v in listas.items()})

dados = {"baseline":65, "itens":itens_out, "combos":combos_out, "mapas":mapas_final,
         "listas":listas,
         "excluir":["acucar","açúcar","arroz","feijao","feijão","sal","tempero","temperos"],
         "unidades":["kg","un","pç","pct","cx","lt","ml","bag","lata","mç","g"]}
out = "/home/user/tatasushireservas/src/lib/cardapio/dados.json"
with open(out,"w",encoding="utf-8") as f:
    json.dump(dados,f,ensure_ascii=False,separators=(",",":"))
import os; print("OK:", os.path.getsize(out)//1024,"KB")
