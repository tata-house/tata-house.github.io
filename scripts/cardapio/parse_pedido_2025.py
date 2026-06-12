# -*- coding: utf-8 -*-
"""Parser unificado das 47 semanas do Pedido 2025 (3 layouts)."""
import openpyxl, unicodedata, re, warnings, json
warnings.filterwarnings("ignore")

def nk(s):
    if s is None: return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+"," ",s).strip().lower()

def clean(s):
    if s is None: return None
    s = re.sub(r"\s+"," ",str(s).replace("￼","")).strip(" .;")
    return s or None

UNID = {"kg":"kg","kgs":"kg","kilo":"kg","kilos":"kg","quilo":"kg","un":"un","uni":"un","unid":"un",
        "unidade":"un","unidades":"un","und":"un","pct":"pct","pcts":"pct","pacote":"pct","pacotes":"pct",
        "bag":"bag","peça":"pç","peca":"pç","pç":"pç","pc":"pç","peças":"pç","pecas":"pç",
        "lt":"lt","litro":"lt","litros":"lt","l":"lt","lata":"lata","latas":"lata","cx":"cx","caixa":"cx",
        "caixas":"cx","maço":"mç","maco":"mç","maços":"mç","macos":"mç","mç":"mç","dz":"dz","ml":"ml",
        "g":"g","gr":"g","balde":"balde","baldes":"balde","galão":"galão","galao":"galão","fardo":"fardo",
        "saco":"saco","sacos":"saco","bandeja":"bandeja","bandejas":"bandeja","pote":"pote","potes":"pote"}

QTD_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*([a-zçãõéíáú]*\.?)", re.I)

def parse_qtd_txt(s):
    """'20 kg'/'20kg'/'01 cx' -> (20.0,'kg')"""
    if s is None: return None
    if isinstance(s,(int,float)): return (float(s), None)
    m = QTD_RE.search(str(s).replace(",","."))
    if not m: return None
    u = UNID.get(nk(m.group(2)), None)
    return (float(m.group(1)), u)

DIA_RE = re.compile(r"^\s*(SEGUNDA|TER[CÇ]A|QUARTA|QUINTA|SEXTA|S[ÁA]BADO|DOMINGO)", re.I)
DIA_IDX = {"SEGUNDA":0,"TERCA":1,"QUARTA":2,"QUINTA":3,"SEXTA":4,"SABADO":5,"DOMINGO":6}
CATS = {"prato principal":"p","guarnição fixa":"gf","guarnicao fixa":"gf","guarnição":"g",
        "guarnicao":"g","guarnições":"g","salada":"s","saladas":"s","sobremesa":"sb","sobremesas":"sb"}

def parse_menu_text(txt, menu):
    """'Prato Principal: X\\nGuarnição: Y...' -> preenche menu"""
    for ln in re.split(r"[\n\r]+", str(txt)):
        if ":" in ln:
            lab, _, val = ln.partition(":")
            ck = CATS.get(nk(lab))
            if ck and clean(val): menu[ck] = clean(val)

def split_item_qty(a):
    """'Bisteca de porco- 20 kg' -> ('Bisteca de porco', 20.0, 'kg')"""
    m = re.match(r"^(.*?)[-–—]\s*(\d.*)$", a)
    if m and len(clean(m.group(1)) or "") >= 3:
        pq = parse_qtd_txt(m.group(2))
        if pq: return clean(m.group(1)), pq[0], pq[1]
    return clean(a), None, None

def parse_pedido_2025(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    nsheets = len(wb.worksheets)
    records = []
    for idx, ws in enumerate(wb.worksheets):
        seq = nsheets - idx
        maxr = ws.max_row or 0
        blocks = []
        for r in range(1, maxr+1):
            v = ws.cell(r,1).value
            if isinstance(v,str):
                m = DIA_RE.match(nk(v).upper())
                if m and len(nk(v)) < 40:
                    blocks.append((r, DIA_IDX[m.group(1).replace("Ç","C").replace("Á","A")]))
        # layout antigo: 1º bloco pode começar na linha 1 sem header (A1=' ')
        if blocks and blocks[0][0] > 3 and any(
            isinstance(ws.cell(r,1).value,str) and "prato principal" in nk(ws.cell(r,1).value)
            for r in range(1, blocks[0][0])):
            blocks.insert(0, (0, (blocks[0][1]-1) % 7))
        for bi,(start, dia) in enumerate(blocks):
            end = blocks[bi+1][0] if bi+1 < len(blocks) else maxr+1
            menu = {"p":None,"gf":None,"g":None,"s":None,"sb":None}
            itens = []
            in_items = False; qcol = None
            for r in range(start+1, end):
                a = ws.cell(r,1).value
                la = nk(a)
                if not in_items:
                    if la in CATS:                       # rótulo em A, valor em B
                        menu[CATS[la]] = clean(ws.cell(r,2).value); continue
                    if "prato principal" in la and ":" in str(a):
                        parse_menu_text(a, menu); continue
                    if la.startswith("item necess"):
                        in_items = True
                        # acha coluna da quantidade no cabeçalho (B ou C)
                        qcol = None
                        for c in (2,3):
                            h = nk(ws.cell(r,c).value)
                            if h.startswith(("qtd","quant","qaunt")): qcol = c; break
                        continue
                    continue
                # zona de itens
                nome = clean(a)
                if not nome: continue
                if nk(nome).startswith("item necess"): continue
                qv = ws.cell(r,qcol).value if qcol else None
                if qv is not None:
                    if "estoque" in nk(qv): continue
                    pq = parse_qtd_txt(qv)
                    if pq:
                        for parte in re.split(r"\s*/\s*", nome):
                            p = clean(parte)
                            if p: itens.append((p, pq[0], pq[1]))
                        continue
                if "estoque" in nk(nome): continue
                for parte in re.split(r"\s*/\s*", str(nome)):
                    n,q,u = split_item_qty(parte)
                    if n and q is not None: itens.append((n,q,u))
            if menu["p"] and itens:
                records.append({"seq":seq,"dia":dia,"menu":menu,"itens":itens})
    return records

if __name__ == "__main__":
    U = "/root/.claude/uploads/28a7abfe-f16f-54cd-9c25-3a5680d3125b/"
    recs = parse_pedido_2025(U+"76d86e48-Planilha_Tata_House_Pedido_2025.xlsx")
    print("dias:", len(recs))
    from collections import Counter
    print("itens/dia médio:", round(sum(len(r["itens"]) for r in recs)/len(recs),1))
    ex = [r for r in recs if r["seq"]==24][:2]
    for r in ex: print(r["menu"], r["itens"][:5])
